from __future__ import annotations

import csv
import io
import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import desc, func, select
from sqlalchemy.orm import Session

from app.models import (
    AdvertisedProductMetric,
    BulkOperationItem,
    BusinessMetric,
    CampaignMetric,
    CostItem,
    ImportBatch,
    InventoryItem,
    ListingItem,
    SearchTermMetric,
    TargetingMetric,
)


REPORT_TYPES = {
    "business": "Business Report",
    "search_terms": "Advertising Search Term Report",
    "advertised_products": "Advertised Product Report",
    "campaigns": "Campaign Report",
    "targeting": "Targeting Report",
    "bulk_operations": "Bulk Operations",
    "inventory": "Inventory Report",
    "costs": "成本表",
    "listings": "Listing 基础信息表",
}

REPORT_INBOX = Path("amazon_reports/inbox")
REPORT_IMPORTED = Path("amazon_reports/imported")
REPORT_FAILED = Path("amazon_reports/failed")
SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}


ALIASES = {
    "sku": ["sku", "childsku", "seller sku", "advertisedsku", "advertised sku", "merchant sku"],
    "asin": ["asin", "childasin", "child asin", "advertisedasin", "advertised asin"],
    "title": ["title", "product name", "productname", "item name"],
    "sessions": ["sessions", "sessions total", "total sessions"],
    "page_views": ["page views", "pageviews", "page views total"],
    "units_ordered": ["units ordered", "unitsordered", "ordered units", "units"],
    "ordered_sales": ["ordered product sales", "orderedproductsales", "sales", "ordered revenue"],
    "date": ["date", "日期"],
    "conversion_rate": ["unit session percentage", "unit session %", "conversion rate", "cvr"],
    "buy_box_percentage": ["buy box percentage", "buy box %"],
    "campaign_name": ["campaign name", "campaign", "广告活动名称"],
    "campaign_id": ["campaign id", "campaignid"],
    "campaign_status": ["campaign status", "campaign state", "state", "status", "广告活动状态"],
    "ad_group_name": ["ad group name", "ad group", "广告组名称"],
    "ad_group_id": ["ad group id", "adgroupid"],
    "targeting": ["targeting", "target", "keyword", "keyword or product targeting", "投放"],
    "match_type": ["match type", "matchtype", "匹配类型"],
    "targeting_status": ["targeting status", "keyword status", "status", "state", "投放状态"],
    "bid": ["bid", "max bid", "竞价"],
    "budget": ["budget", "daily budget", "campaign budget", "预算", "每日预算"],
    "search_term": ["customer search term", "search term", "customersearchterm", "客户搜索词", "搜索词"],
    "impressions": ["impressions", "展示量"],
    "clicks": ["clicks", "点击量"],
    "spend": ["spend", "cost", "ad spend", "花费"],
    "ad_sales": ["7 day total sales", "14 day total sales", "sales", "total sales", "7天总销售额", "14天总销售额", "销售额"],
    "orders": ["7 day total orders", "14 day total orders", "orders", "total orders", "7天总订单数", "14天总订单数", "订单数"],
    "acos": ["acos", "total advertising cost of sales acos", "advertising cost of sales", "广告投入产出比", "总广告投入产出比"],
    "purchase_cost": ["purchase cost", "product cost", "unit cost", "采购成本"],
    "first_leg_shipping": ["first leg shipping", "shipping cost", "头程", "头程运费"],
    "packaging_cost": ["packaging cost", "package cost", "包装成本"],
    "fba_fee": ["fba fee", "fulfillment fee", "配送费"],
    "referral_fee_rate": ["referral fee rate", "commission rate", "佣金比例"],
    "other_cost": ["other cost", "misc cost", "其他成本"],
    "target_margin": ["target margin", "目标利润率"],
    "product_name": ["product name", "product", "产品名称"],
    "bullet_1": ["bullet 1", "bullet_1", "bullet point 1"],
    "bullet_2": ["bullet 2", "bullet_2", "bullet point 2"],
    "bullet_3": ["bullet 3", "bullet_3", "bullet point 3"],
    "bullet_4": ["bullet 4", "bullet_4", "bullet point 4"],
    "bullet_5": ["bullet 5", "bullet_5", "bullet point 5"],
    "price": ["price", "售价", "sale price"],
    "coupon": ["coupon", "优惠券"],
    "main_image_url": ["main image url", "image url", "main_image_url"],
    "record_type": ["record type", "entity", "entity type", "product"],
    "operation": ["operation"],
    "entity_id": ["entity id", "keyword id", "target id", "ad id"],
    "entity_status": ["entity status", "status", "state"],
    "keyword_text": ["keyword text", "keyword", "keyword or product targeting"],
    "targeting_expression": ["targeting expression", "product targeting expression", "targeting"],
    "fnsku": ["fnsku"],
    "available": ["available", "afn fulfillable quantity", "sellable on hand quantity"],
    "inbound": ["inbound", "afn inbound working quantity", "inbound quantity"],
    "reserved": ["reserved", "afn reserved quantity"],
    "days_of_supply": ["days of supply", "estimated days of supply"],
}


@dataclass
class SkuSummary:
    sku: str
    asin: str
    title: str
    sales: float
    units: float
    sessions: float
    conversion_rate: float | None
    ad_spend: float
    ad_sales: float
    ad_orders: float
    acos: float | None
    tacos: float | None
    unit_cost: float
    estimated_profit: float
    margin: float | None
    tags: list[str]
    recommendations: list[str]


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").strip().lower())


def _alias_map(headers: list[str]) -> dict[str, str]:
    normalized = {normalize_header(header): header for header in headers}
    result = {}
    for canonical, aliases in ALIASES.items():
        for alias in aliases:
            normalized_alias = normalize_header(alias)
            hit = normalized.get(normalized_alias)
            if hit:
                result[canonical] = hit
                break
            fuzzy_hit = next(
                (
                    original
                    for normalized_header, original in normalized.items()
                    if normalized_alias and normalized_alias in normalized_header
                ),
                None,
            )
            if fuzzy_hit:
                result[canonical] = fuzzy_hit
                break
    return result


def _get(row: dict, alias: dict[str, str], key: str, default: str = ""):
    source = alias.get(key)
    return row.get(source, default) if source else default


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "").replace("USD", "")
    is_percent = "%" in text
    text = text.replace("%", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if is_percent and number > 1:
        return number / 100
    return number


def _json(row: dict) -> str:
    return json.dumps(row, ensure_ascii=False, default=str)


def _is_empty_cell(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _header_score(values: list[Any]) -> int:
    headers = [str(value or "").strip() for value in values]
    alias_hits = len(_alias_map(headers))
    non_empty = sum(1 for value in headers if value)
    long_text_hits = sum(1 for value in headers if len(str(value)) > 80)
    return alias_hits * 3 + min(non_empty, 6) - long_text_hits * 2


def _dict_rows_from_matrix(rows: list[tuple[Any, ...]], sheet_name: str | None = None) -> list[dict]:
    if not rows:
        return []

    best_index = 0
    best_score = -1
    scan_limit = min(len(rows), 80)
    for index in range(scan_limit):
        values = list(rows[index])
        if not any(not _is_empty_cell(value) for value in values):
            continue
        score = _header_score(values)
        if score > best_score:
            best_score = score
            best_index = index

    if best_score < 6:
        best_index = 0

    raw_headers = [str(value or "").strip() for value in rows[best_index]]
    headers = []
    seen: dict[str, int] = defaultdict(int)
    for index, header in enumerate(raw_headers):
        cleaned = header or f"column_{index + 1}"
        seen[cleaned] += 1
        if seen[cleaned] > 1:
            cleaned = f"{cleaned}_{seen[cleaned]}"
        headers.append(cleaned)

    parsed = []
    for values in rows[best_index + 1 :]:
        if not any(not _is_empty_cell(value) for value in values):
            continue
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if sheet_name:
            row["_sheet_name"] = sheet_name
        parsed.append(row)
    return parsed


def parse_tabular_file(file_name: str, content: bytes) -> list[dict]:
    lower = file_name.lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        text = content.decode("utf-8-sig", errors="ignore")
        dialect = csv.excel_tab if lower.endswith(".tsv") else csv.excel
        matrix = list(csv.reader(io.StringIO(text), dialect=dialect))
        return _dict_rows_from_matrix([tuple(row) for row in matrix])
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        parsed = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            sheet_rows = _dict_rows_from_matrix(rows, sheet.title)
            if sheet_rows and len(_alias_map(list(sheet_rows[0].keys()))) >= 2:
                parsed.extend(sheet_rows)
        return parsed
    raise ValueError("只支持 CSV、TSV、XLSX、XLSM 文件。")


def import_report(db: Session, report_type: str, file_name: str, content: bytes) -> ImportBatch:
    if report_type not in REPORT_TYPES:
        raise ValueError("未知报表类型")
    batch = ImportBatch(report_type=report_type, file_name=file_name, status="running")
    db.add(batch)
    db.commit()
    db.refresh(batch)
    try:
        rows = parse_tabular_file(file_name, content)
        if not rows:
            raise ValueError("文件没有可读取的数据行。")
        for row in rows:
            alias = _alias_map(list(row.keys()))
            _insert_row(db, batch.id, report_type, row, alias)
        batch.status = "success"
        batch.row_count = len(rows)
    except Exception as exc:
        db.rollback()
        batch = db.get(ImportBatch, batch.id)
        batch.status = "failed"
        batch.error_message = str(exc)
    db.commit()
    db.refresh(batch)
    return batch


def ensure_report_dirs() -> None:
    for path in (REPORT_INBOX, REPORT_IMPORTED, REPORT_FAILED):
        path.mkdir(parents=True, exist_ok=True)


def infer_report_type(file_name: str, rows: list[dict]) -> str:
    name = file_name.lower()
    headers = set()
    for row in rows[:20]:
        headers.update(normalize_header(header) for header in row.keys())
    if any(term in name for term in ["bulk", "bulk_operations", "bulk-operations", "批量"]):
        return "bulk_operations"
    if any(term in name for term in ["campaign", "campaigns", "广告活动"]):
        return "campaigns"
    if any(term in name for term in ["targeting", "targeting_report", "投放"]):
        return "targeting"
    if any(term in name for term in ["inventory", "fba_inventory", "库存"]):
        return "inventory"
    if any(term in name for term in ["search_term", "search-term", "searchterm", "搜索词"]):
        return "search_terms"
    if any(term in name for term in ["advertised_product", "advertised-product", "ad_products", "广告商品"]):
        return "advertised_products"
    if any(term in name for term in ["cost", "成本"]):
        return "costs"
    if any(term in name for term in ["listing", "listings", "链接", "listing表"]):
        return "listings"
    if any(term in name for term in ["business", "sales_traffic", "sales-and-traffic", "业务"]):
        return "business"

    if normalize_header("Customer Search Term") in headers or normalize_header("客户搜索词") in headers:
        return "search_terms"
    if normalize_header("Entity") in headers or normalize_header("Record Type") in headers:
        return "bulk_operations"
    if normalize_header("Campaign ID") in headers and normalize_header("Budget") in headers:
        return "campaigns"
    if normalize_header("Match Type") in headers and normalize_header("Bid") in headers:
        return "targeting"
    if normalize_header("fnsku") in headers or normalize_header("afn fulfillable quantity") in headers:
        return "inventory"
    if normalize_header("Advertised SKU") in headers or normalize_header("Advertised ASIN") in headers:
        return "advertised_products"
    if normalize_header("purchase_cost") in headers or normalize_header("采购成本") in headers:
        return "costs"
    if normalize_header("bullet_1") in headers or normalize_header("bullet point 1") in headers:
        return "listings"
    if normalize_header("Ordered Product Sales") in headers or normalize_header("已订购商品销售额") in headers:
        return "business"
    raise ValueError("无法自动识别报表类型，请在网页手动选择类型上传。")


def scan_inbox(db: Session) -> list[dict]:
    ensure_report_dirs()
    results = []
    for path in sorted(REPORT_INBOX.iterdir()):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        try:
            content = path.read_bytes()
            rows = parse_tabular_file(path.name, content)
            report_type = infer_report_type(path.name, rows)
            batch = import_report(db, report_type, path.name, content)
            target_dir = REPORT_IMPORTED if batch.status == "success" else REPORT_FAILED
            destination = target_dir / f"{path.stem}-{timestamp}{path.suffix}"
            shutil.move(str(path), destination)
            results.append(
                {
                    "file_name": path.name,
                    "report_type": report_type,
                    "status": batch.status,
                    "row_count": batch.row_count,
                    "error_message": batch.error_message,
                    "destination": str(destination),
                }
            )
        except Exception as exc:
            db.rollback()
            destination = REPORT_FAILED / f"{path.stem}-{timestamp}{path.suffix}"
            shutil.move(str(path), destination)
            results.append(
                {
                    "file_name": path.name,
                    "report_type": "",
                    "status": "failed",
                    "row_count": 0,
                    "error_message": str(exc),
                    "destination": str(destination),
                }
            )
    return results


def _insert_row(db: Session, batch_id: int, report_type: str, row: dict, alias: dict[str, str]) -> None:
    common = {"batch_id": batch_id, "raw_json": _json(row)}
    if report_type == "business":
        db.add(
            BusinessMetric(
                **common,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                title=str(_get(row, alias, "title")).strip() or None,
                sessions=to_float(_get(row, alias, "sessions")),
                page_views=to_float(_get(row, alias, "page_views")),
                units_ordered=to_float(_get(row, alias, "units_ordered")),
                ordered_sales=to_float(_get(row, alias, "ordered_sales")),
                conversion_rate=to_float(_get(row, alias, "conversion_rate")),
                buy_box_percentage=to_float(_get(row, alias, "buy_box_percentage")),
            )
        )
    elif report_type == "search_terms":
        db.add(
            SearchTermMetric(
                **common,
                campaign_name=str(_get(row, alias, "campaign_name")).strip() or None,
                ad_group_name=str(_get(row, alias, "ad_group_name")).strip() or None,
                targeting=str(_get(row, alias, "targeting")).strip() or None,
                search_term=str(_get(row, alias, "search_term")).strip() or None,
                impressions=to_float(_get(row, alias, "impressions")),
                clicks=to_float(_get(row, alias, "clicks")),
                spend=to_float(_get(row, alias, "spend")),
                sales=to_float(_get(row, alias, "ad_sales")),
                orders=to_float(_get(row, alias, "orders")),
                acos=to_float(_get(row, alias, "acos")),
            )
        )
    elif report_type == "advertised_products":
        db.add(
            AdvertisedProductMetric(
                **common,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                campaign_name=str(_get(row, alias, "campaign_name")).strip() or None,
                impressions=to_float(_get(row, alias, "impressions")),
                clicks=to_float(_get(row, alias, "clicks")),
                spend=to_float(_get(row, alias, "spend")),
                sales=to_float(_get(row, alias, "ad_sales")),
                orders=to_float(_get(row, alias, "orders")),
                acos=to_float(_get(row, alias, "acos")),
            )
        )
    elif report_type == "campaigns":
        db.add(
            CampaignMetric(
                **common,
                campaign_name=str(_get(row, alias, "campaign_name")).strip() or None,
                campaign_id=str(_get(row, alias, "campaign_id")).strip() or None,
                campaign_status=str(_get(row, alias, "campaign_status")).strip() or None,
                impressions=to_float(_get(row, alias, "impressions")),
                clicks=to_float(_get(row, alias, "clicks")),
                spend=to_float(_get(row, alias, "spend")),
                sales=to_float(_get(row, alias, "ad_sales")),
                orders=to_float(_get(row, alias, "orders")),
                acos=to_float(_get(row, alias, "acos")),
                budget=to_float(_get(row, alias, "budget")),
            )
        )
    elif report_type == "targeting":
        db.add(
            TargetingMetric(
                **common,
                campaign_name=str(_get(row, alias, "campaign_name")).strip() or None,
                ad_group_name=str(_get(row, alias, "ad_group_name")).strip() or None,
                targeting=str(_get(row, alias, "targeting")).strip() or None,
                match_type=str(_get(row, alias, "match_type")).strip() or None,
                status=str(_get(row, alias, "targeting_status")).strip() or None,
                bid=to_float(_get(row, alias, "bid")),
                impressions=to_float(_get(row, alias, "impressions")),
                clicks=to_float(_get(row, alias, "clicks")),
                spend=to_float(_get(row, alias, "spend")),
                sales=to_float(_get(row, alias, "ad_sales")),
                orders=to_float(_get(row, alias, "orders")),
                acos=to_float(_get(row, alias, "acos")),
            )
        )
    elif report_type == "bulk_operations":
        db.add(
            BulkOperationItem(
                **common,
                record_type=str(_get(row, alias, "record_type")).strip() or None,
                operation=str(_get(row, alias, "operation")).strip() or None,
                campaign_name=str(_get(row, alias, "campaign_name")).strip() or None,
                campaign_id=str(_get(row, alias, "campaign_id")).strip() or None,
                campaign_status=str(_get(row, alias, "campaign_status")).strip() or None,
                ad_group_name=str(_get(row, alias, "ad_group_name")).strip() or None,
                ad_group_id=str(_get(row, alias, "ad_group_id")).strip() or None,
                entity_id=str(_get(row, alias, "entity_id")).strip() or None,
                entity_status=str(_get(row, alias, "entity_status")).strip() or None,
                keyword_text=str(_get(row, alias, "keyword_text")).strip() or None,
                match_type=str(_get(row, alias, "match_type")).strip() or None,
                targeting_expression=str(_get(row, alias, "targeting_expression")).strip() or None,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                bid=to_float(_get(row, alias, "bid")),
                budget=to_float(_get(row, alias, "budget")),
            )
        )
    elif report_type == "inventory":
        db.add(
            InventoryItem(
                **common,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                fnsku=str(_get(row, alias, "fnsku")).strip() or None,
                product_name=str(_get(row, alias, "product_name")).strip() or None,
                available=to_float(_get(row, alias, "available")),
                inbound=to_float(_get(row, alias, "inbound")),
                reserved=to_float(_get(row, alias, "reserved")),
                days_of_supply=to_float(_get(row, alias, "days_of_supply")),
            )
        )
    elif report_type == "costs":
        db.add(
            CostItem(
                **common,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                product_name=str(_get(row, alias, "product_name")).strip() or None,
                purchase_cost=to_float(_get(row, alias, "purchase_cost")),
                first_leg_shipping=to_float(_get(row, alias, "first_leg_shipping")),
                packaging_cost=to_float(_get(row, alias, "packaging_cost")),
                fba_fee=to_float(_get(row, alias, "fba_fee")),
                referral_fee_rate=to_float(_get(row, alias, "referral_fee_rate")),
                other_cost=to_float(_get(row, alias, "other_cost")),
                target_margin=to_float(_get(row, alias, "target_margin")),
            )
        )
    elif report_type == "listings":
        db.add(
            ListingItem(
                **common,
                sku=str(_get(row, alias, "sku")).strip() or None,
                asin=str(_get(row, alias, "asin")).strip() or None,
                title=str(_get(row, alias, "title")).strip() or None,
                bullet_1=str(_get(row, alias, "bullet_1")).strip() or None,
                bullet_2=str(_get(row, alias, "bullet_2")).strip() or None,
                bullet_3=str(_get(row, alias, "bullet_3")).strip() or None,
                bullet_4=str(_get(row, alias, "bullet_4")).strip() or None,
                bullet_5=str(_get(row, alias, "bullet_5")).strip() or None,
                price=to_float(_get(row, alias, "price")),
                coupon=str(_get(row, alias, "coupon")).strip() or None,
                main_image_url=str(_get(row, alias, "main_image_url")).strip() or None,
            )
        )


def latest_batches(db: Session, limit: int = 10, offset: int = 0) -> list[ImportBatch]:
    return (
        db.execute(select(ImportBatch).order_by(desc(ImportBatch.created_at)).offset(offset).limit(limit))
        .scalars()
        .all()
    )


def count_batches(db: Session) -> int:
    return db.execute(select(func.count(ImportBatch.id))).scalar_one()


def build_sku_dashboard(db: Session) -> list[SkuSummary]:
    business_rows = db.execute(select(BusinessMetric)).scalars().all()
    ad_rows = db.execute(select(AdvertisedProductMetric)).scalars().all()
    costs = db.execute(select(CostItem).order_by(desc(CostItem.created_at))).scalars().all()
    listings = db.execute(select(ListingItem).order_by(desc(ListingItem.created_at))).scalars().all()

    by_sku: dict[str, dict] = defaultdict(lambda: defaultdict(float))
    meta: dict[str, dict] = defaultdict(dict)

    for row in business_rows:
        key = row.sku or row.asin or "unknown"
        by_sku[key]["sales"] += row.ordered_sales or 0
        by_sku[key]["units"] += row.units_ordered or 0
        by_sku[key]["sessions"] += row.sessions or 0
        meta[key].update({"sku": row.sku or key, "asin": row.asin or "", "title": row.title or meta[key].get("title", "")})

    for row in ad_rows:
        key = row.sku or row.asin or "unknown"
        by_sku[key]["ad_spend"] += row.spend or 0
        by_sku[key]["ad_sales"] += row.sales or 0
        by_sku[key]["ad_orders"] += row.orders or 0
        meta[key].update({"sku": row.sku or key, "asin": row.asin or meta[key].get("asin", "")})

    cost_by_key = {}
    for cost in costs:
        key = cost.sku or cost.asin
        if key and key not in cost_by_key:
            cost_by_key[key] = cost
    for listing in listings:
        key = listing.sku or listing.asin
        if key:
            meta[key].update({"sku": listing.sku or key, "asin": listing.asin or meta[key].get("asin", ""), "title": listing.title or meta[key].get("title", "")})

    summaries = []
    for key, values in by_sku.items():
        sales = values["sales"]
        units = values["units"]
        sessions = values["sessions"]
        ad_spend = values["ad_spend"]
        ad_sales = values["ad_sales"]
        ad_orders = values["ad_orders"]
        cost = cost_by_key.get(key) or cost_by_key.get(meta[key].get("asin", ""))
        unit_cost = 0.0
        referral_rate = 0.15
        target_margin = 0.25
        if cost:
            unit_cost = sum(
                value or 0
                for value in [cost.purchase_cost, cost.first_leg_shipping, cost.packaging_cost, cost.fba_fee, cost.other_cost]
            )
            referral_rate = cost.referral_fee_rate if cost.referral_fee_rate is not None else referral_rate
            if referral_rate > 1:
                referral_rate = referral_rate / 100
            target_margin = cost.target_margin if cost.target_margin is not None else target_margin
            if target_margin > 1:
                target_margin = target_margin / 100
        estimated_profit = sales - ad_spend - sales * referral_rate - units * unit_cost
        margin = estimated_profit / sales if sales else None
        conversion_rate = units / sessions if sessions else None
        acos = ad_spend / ad_sales if ad_sales else None
        tacos = ad_spend / sales if sales else None
        tags, recommendations = _sku_tags_and_recommendations(
            sales, units, sessions, ad_spend, ad_sales, conversion_rate, acos, tacos, margin, target_margin
        )
        summaries.append(
            SkuSummary(
                sku=meta[key].get("sku") or key,
                asin=meta[key].get("asin") or "",
                title=meta[key].get("title") or "",
                sales=sales,
                units=units,
                sessions=sessions,
                conversion_rate=conversion_rate,
                ad_spend=ad_spend,
                ad_sales=ad_sales,
                ad_orders=ad_orders,
                acos=acos,
                tacos=tacos,
                unit_cost=unit_cost,
                estimated_profit=estimated_profit,
                margin=margin,
                tags=tags,
                recommendations=recommendations,
            )
        )
    return sorted(summaries, key=lambda item: item.estimated_profit, reverse=True)


def _sku_tags_and_recommendations(
    sales: float,
    units: float,
    sessions: float,
    ad_spend: float,
    ad_sales: float,
    conversion_rate: float | None,
    acos: float | None,
    tacos: float | None,
    margin: float | None,
    target_margin: float,
) -> tuple[list[str], list[str]]:
    tags = []
    recs = []
    if margin is not None and margin >= target_margin:
        tags.append("利润健康")
    if margin is not None and margin < target_margin:
        tags.append("利润偏低")
        recs.append("检查售价、广告花费和成本；优先保留高转化词，压缩低效广告。")
    if conversion_rate is not None and conversion_rate < 0.05 and sessions >= 100:
        tags.append("转化偏低")
        recs.append("优化主图、尺寸图、适配说明和价格锚点。")
    if conversion_rate is not None and conversion_rate >= 0.12 and sessions < 300:
        tags.append("高转化低流量")
        recs.append("增加精准词广告预算，扩展相关长尾词。")
    if acos is not None and acos > 0.45 and ad_spend >= 20:
        tags.append("广告亏损风险")
        recs.append("降低 broad/auto 出价，检查浪费搜索词。")
    if acos is not None and acos < 0.25 and ad_sales >= 50:
        tags.append("广告可放大")
        recs.append("把高转化词拆 exact 活动，适度加预算。")
    if sales > 0 and ad_spend == 0:
        tags.append("自然销售")
        recs.append("可小预算测试精准词，确认是否能放量。")
    if units >= 10 and margin is not None and margin > target_margin:
        tags.append("可测试多件装")
        recs.append("评估 3-pack/5-pack，提高客单价和广告承受能力。")
    return tags or ["待观察"], recs or ["继续积累数据，等待更多销量和广告表现。"]


def build_ad_actions(db: Session) -> list[dict]:
    rows = db.execute(select(SearchTermMetric).order_by(desc(SearchTermMetric.spend))).scalars().all()
    advertised_rows = db.execute(select(AdvertisedProductMetric)).scalars().all()
    listings = db.execute(select(ListingItem).order_by(desc(ListingItem.created_at))).scalars().all()
    business_rows = db.execute(select(BusinessMetric).order_by(desc(BusinessMetric.created_at))).scalars().all()

    title_by_key = {}
    for item in [*listings, *business_rows]:
        for key in [getattr(item, "sku", None), getattr(item, "asin", None)]:
            if key and key not in title_by_key:
                title_by_key[key] = getattr(item, "title", None) or getattr(item, "product_name", None) or ""

    ad_candidates_by_pair: dict[tuple[str, str], dict[tuple[str, str], dict]] = defaultdict(dict)
    ad_candidates_by_campaign: dict[str, dict[tuple[str, str], dict]] = defaultdict(dict)
    for ad in advertised_rows:
        campaign = ad.campaign_name or ""
        ad_group = _raw_text(ad, "Ad Group Name", "广告组名称")
        sku = ad.sku or ""
        asin = ad.asin or ""
        if not (campaign or ad_group or sku or asin):
            continue
        key = (asin, sku)
        candidate = {
            "asin": asin,
            "sku": sku,
            "title": title_by_key.get(sku) or title_by_key.get(asin) or "",
            "ad_group": ad_group,
            "campaign": campaign,
        }
        ad_candidates_by_pair[(campaign, ad_group)][key] = candidate
        ad_candidates_by_campaign[campaign][key] = candidate

    actions = []
    merged_rows: dict[tuple[str, str, str, str, str], dict] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        campaign = row.campaign_name or ""
        ad_group = row.ad_group_name or ""
        targeting = row.targeting or ""
        search_term = row.search_term or row.targeting or ""
        match_type = _raw_text(row, "匹配类型", "Match Type")
        key = (campaign, ad_group, targeting, search_term, match_type)
        merged = merged_rows[key]
        merged["campaign"] = campaign
        merged["ad_group"] = ad_group
        merged["targeting"] = targeting
        merged["search_term"] = search_term
        merged["match_type"] = match_type
        merged["clicks"] += row.clicks or 0
        merged["spend"] += row.spend or 0
        merged["sales"] += row.sales or 0
        merged["orders"] += row.orders or 0
        merged["impressions"] += row.impressions or 0

    for row in merged_rows.values():
        clicks = row["clicks"]
        spend = row["spend"]
        sales = row["sales"]
        orders = row["orders"]
        impressions = row["impressions"]
        ctr = clicks / impressions if impressions else None
        acos = spend / sales if sales else None
        priority = None
        action = None
        reason = None
        if clicks >= 12 and orders == 0 and spend >= 10:
            priority = "P0"
            action = "否定 exact 或大幅降价"
            reason = "点击和花费已经足够，但没有订单。"
        elif orders >= 2 and acos is not None and acos <= 0.25:
            priority = "P1"
            action = "加入 exact 活动并适度加预算"
            reason = "搜索词有订单且 ACOS 较低。"
        elif orders >= 1 and sales > 0 and spend / sales <= 0.35:
            priority = "P1"
            action = "保留并观察，可测试 phrase/exact"
            reason = "已有转化，广告效率可接受。"
        elif ctr is not None and ctr < 0.002 and impressions >= 2000:
            priority = "P2"
            action = "检查主图/标题相关性或降低出价"
            reason = "曝光高但点击率低。"
        if priority:
            campaign = row["campaign"]
            ad_group = row["ad_group"]
            candidates = list(ad_candidates_by_pair.get((campaign, ad_group), {}).values())
            match_type = "campaign+ad_group"
            if not candidates:
                candidates = list(ad_candidates_by_campaign.get(campaign, {}).values())
                match_type = "campaign"
            if not candidates:
                candidates = [{"asin": "", "sku": "", "title": "", "campaign": campaign, "ad_group": ad_group}]
                match_type = "unmatched"
            actions.append(
                {
                    "priority": priority,
                    "action": action,
                    "reason": reason,
                    "search_term": row["search_term"],
                    "targeting": row["targeting"],
                    "match_type": row["match_type"],
                    "campaign": campaign,
                    "ad_group": ad_group,
                    "linked_products": candidates,
                    "link_method": match_type,
                    "clicks": clicks,
                    "spend": spend,
                    "sales": sales,
                    "orders": orders,
                    "acos": acos,
                }
            )
    priority_order = {"P0": 0, "P1": 1, "P2": 2}
    return sorted(actions, key=lambda item: (priority_order[item["priority"]], -item["spend"]))[:100]


def build_ad_action_groups(db: Session) -> list[dict]:
    actions = build_ad_actions(db)
    priority_order = {"P0": 0, "P1": 1, "P2": 2}
    grouped: dict[tuple[str, str], dict] = {}
    unknown_index = 0
    for action in actions:
        linked_products = action.get("linked_products") or []
        for product in linked_products:
            asin = product.get("asin") or ""
            sku = product.get("sku") or ""
            if not asin and not sku:
                unknown_index += 1
                key = (f"未关联-{unknown_index}", "")
            else:
                key = (asin, sku)
            group = grouped.setdefault(
                key,
                {
                    "asin": asin,
                    "sku": sku,
                    "title": product.get("title") or "",
                    "actions": [],
                    "total_spend": 0.0,
                    "total_clicks": 0.0,
                    "total_sales": 0.0,
                    "total_orders": 0.0,
                    "highest_priority": "P2",
                },
            )
            group["actions"].append(action)
            group["total_spend"] += action.get("spend") or 0
            group["total_clicks"] += action.get("clicks") or 0
            group["total_sales"] += action.get("sales") or 0
            group["total_orders"] += action.get("orders") or 0
            if priority_order[action["priority"]] < priority_order[group["highest_priority"]]:
                group["highest_priority"] = action["priority"]
    for group in grouped.values():
        group["acos"] = group["total_spend"] / group["total_sales"] if group["total_sales"] else None
        group["actions"] = sorted(group["actions"], key=lambda item: (priority_order[item["priority"]], -item["spend"]))
    return sorted(grouped.values(), key=lambda item: (priority_order[item["highest_priority"]], -item["total_spend"]))


def _priority_rank(priority: str) -> int:
    return {"P0": 0, "P1": 1, "P2": 2}.get(priority, 3)


def _latest_success_batches_by_type(db: Session) -> dict[str, ImportBatch]:
    batches = (
        db.execute(select(ImportBatch).where(ImportBatch.status == "success").order_by(desc(ImportBatch.created_at)))
        .scalars()
        .all()
    )
    latest_by_type = {}
    for batch in batches:
        latest_by_type.setdefault(batch.report_type, batch)
    return latest_by_type


def _short_object(asin: str | None, sku: str | None) -> str:
    if asin and sku and asin != sku:
        return f"{asin} / {sku}"
    return asin or sku or "未关联商品"


def build_operations_cockpit(db: Session) -> dict:
    sku_rows = build_sku_dashboard(db)
    ad_groups = build_ad_action_groups(db)
    listing_audits = build_listing_audits(db)
    latest_by_type = _latest_success_batches_by_type(db)

    total_sales = sum(row.sales for row in sku_rows)
    total_ad_spend = sum(row.ad_spend for row in sku_rows)
    total_profit = sum(row.estimated_profit for row in sku_rows)
    p0_ad_groups = sum(1 for group in ad_groups if group.get("highest_priority") == "P0")
    p1_ad_groups = sum(1 for group in ad_groups if group.get("highest_priority") == "P1")
    listing_issue_count = sum(
        1
        for audit in listing_audits
        if audit["issues"] and not audit["issues"][0].startswith("基础信息完整")
    )
    totals = {
        "sku_count": len(sku_rows),
        "sales": total_sales,
        "ad_spend": total_ad_spend,
        "profit": total_profit,
        "tacos": total_ad_spend / total_sales if total_sales else None,
        "margin": total_profit / total_sales if total_sales else None,
        "p0_ad_groups": p0_ad_groups,
        "p1_ad_groups": p1_ad_groups,
        "listing_issue_count": listing_issue_count,
    }

    required_reports = [
        ("business", "Business Report", "销量、Sessions、转化率，是判断放量和转化问题的底座。"),
        ("advertised_products", "Advertised Product Report", "把广告花费落到 SKU/ASIN，识别广告拖累和可放量商品。"),
        ("search_terms", "Search Term Report", "生成否定词、exact 加词和降价动作。"),
        ("bulk_operations", "Bulk Operations", "同步广告活动状态，避免建议已经关闭的活动。"),
        ("campaigns", "Campaign Report", "看活动预算、花费、ACOS 和预算是否限制放量。"),
        ("targeting", "Targeting Report", "看 exact / phrase / broad / ASIN 定投的表现。"),
        ("costs", "成本表", "计算真实利润，不然只能看销售额和广告效率。"),
        ("inventory", "Inventory Report", "判断断货风险和库存压力。"),
        ("listings", "Listing 表", "做标题、五点、图片资料的基础体检。"),
    ]
    data_health = []
    for key, label, why in required_reports:
        batch = latest_by_type.get(key)
        data_health.append(
            {
                "key": key,
                "label": label,
                "ok": bool(batch),
                "why": why,
                "file_name": batch.file_name if batch else "",
                "row_count": batch.row_count if batch else 0,
                "created_at": batch.created_at if batch else None,
            }
        )

    todos = []
    for group in ad_groups[:16]:
        p0_count = sum(1 for action in group["actions"] if action["priority"] == "P0")
        p1_count = sum(1 for action in group["actions"] if action["priority"] == "P1")
        if p0_count:
            priority = "P0"
            next_action = "先处理无转化浪费词：按活动名、广告组、搜索词在广告后台否定 exact；如果是 broad/auto 流量，先降 bid 20%-30%。"
        else:
            priority = "P1"
            next_action = "把有订单且 ACOS 合理的词拆到 exact 活动，单独给预算和 bid，避免被 broad/auto 混在一起。"
        todos.append(
            {
                "priority": priority,
                "area": "广告",
                "object": _short_object(group.get("asin"), group.get("sku")),
                "evidence": f"{len(group['actions'])} 个广告动作，P0 {p0_count} 个，P1 {p1_count} 个，花费 ${group['total_spend']:.2f}，销售 ${group['total_sales']:.2f}。",
                "next_action": next_action,
                "href": "/operations/ad-actions",
            }
        )

    for row in sku_rows:
        label = _short_object(row.asin, row.sku)
        if row.sales > 0 and row.margin is not None and row.margin < 0.15:
            todos.append(
                {
                    "priority": "P0" if row.ad_spend >= 20 else "P1",
                    "area": "利润",
                    "object": label,
                    "evidence": f"销售 ${row.sales:.2f}，预估利润 ${row.estimated_profit:.2f}，利润率 {(row.margin or 0) * 100:.1f}%。",
                    "next_action": "先核对成本和售价；如果成本无误，压缩低效广告词，保留高转化词，并评估涨价或改多件装拉高客单价。",
                    "href": "/operations/dashboard",
                }
            )
        if row.sessions >= 100 and (row.conversion_rate or 0) < 0.05:
            todos.append(
                {
                    "priority": "P1",
                    "area": "转化",
                    "object": label,
                    "evidence": f"Sessions {row.sessions:.0f}，CVR {(row.conversion_rate or 0) * 100:.1f}%，流量够但转化偏低。",
                    "next_action": "优先查主图、尺寸/兼容图、价格锚点、coupon、评论星级；广告先别盲目加预算，等转化修复后再放量。",
                    "href": "/operations/dashboard",
                }
            )
        if row.sales >= 50 and (row.margin or 0) >= 0.25 and (row.conversion_rate or 0) >= 0.12 and (row.tacos or 0) <= 0.18:
            todos.append(
                {
                    "priority": "P1",
                    "area": "放量",
                    "object": label,
                    "evidence": f"利润率 {(row.margin or 0) * 100:.1f}%，CVR {(row.conversion_rate or 0) * 100:.1f}%，TACOS {(row.tacos or 0) * 100:.1f}%。",
                    "next_action": "复制高转化搜索词到 exact，预算小幅增加 20%-30%；同时测试 3-pack/5-pack 或数量折扣，提高广告承受能力。",
                    "href": "/operations/dashboard",
                }
            )
        if row.units >= 10 and (row.margin or 0) >= 0.25:
            todos.append(
                {
                    "priority": "P2",
                    "area": "客单价",
                    "object": label,
                    "evidence": f"销量 {row.units:.0f} 件，利润率 {(row.margin or 0) * 100:.1f}%，具备测试多件装条件。",
                    "next_action": "做 3-pack/5-pack 小批量 FBM 测试，主图突出数量和适配场景；广告先用原 ASIN 高转化词小预算验证。",
                    "href": "/operations/dashboard",
                }
            )

    for audit in listing_audits[:20]:
        listing = audit["listing"]
        issues = [issue for issue in audit["issues"] if not issue.startswith("基础信息完整")]
        if not issues:
            continue
        todos.append(
            {
                "priority": "P1",
                "area": "Listing",
                "object": _short_object(listing.asin, listing.sku),
                "evidence": issues[0],
                "next_action": "先补核心规格、适配词、replacement/compatible/pack 等高意图词，再检查五点是否覆盖尺寸、材质、用途和售后。",
                "href": "/operations/listing-audit",
            }
        )

    for item in data_health:
        if item["ok"]:
            continue
        priority = "P0" if item["key"] in {"costs", "business", "advertised_products", "search_terms"} else "P2"
        todos.append(
            {
                "priority": priority,
                "area": "数据",
                "object": item["label"],
                "evidence": f"缺少 {item['label']}，{item['why']}",
                "next_action": "补上传这个报表后再看对应模块，避免系统基于不完整信息给出偏差建议。",
                "href": "/imports",
            }
        )

    todos = sorted(todos, key=lambda item: (_priority_rank(item["priority"]), item["area"], item["object"]))[:40]
    today_focus = []
    seen_focus = set()
    for item in todos:
        object_root = item["object"].split("/")[0].strip()
        focus_key = (item["area"], object_root or item["object"])
        if focus_key in seen_focus:
            continue
        seen_focus.add(focus_key)
        today_focus.append(item)
        if len(today_focus) >= 5:
            break
    scale_candidates = [
        row
        for row in sku_rows
        if row.sales >= 50 and (row.margin or 0) >= 0.25 and (row.conversion_rate or 0) >= 0.1
    ][:8]
    risk_skus = [
        row
        for row in sku_rows
        if (row.sales > 0 and row.margin is not None and row.margin < 0.15)
        or ((row.acos or 0) > 0.45 and row.ad_spend >= 20)
        or (row.sessions >= 100 and (row.conversion_rate or 0) < 0.05)
    ][:8]

    return {
        "totals": totals,
        "todos": todos,
        "today_focus": today_focus,
        "data_health": data_health,
        "scale_candidates": scale_candidates,
        "risk_skus": risk_skus,
        "ad_groups": ad_groups[:5],
    }


def build_listing_audits(db: Session) -> list[dict]:
    listings = db.execute(select(ListingItem).order_by(desc(ListingItem.created_at))).scalars().all()
    audits = []
    seen = set()
    for item in listings:
        key = item.sku or item.asin
        if not key or key in seen:
            continue
        seen.add(key)
        issues = []
        title = item.title or ""
        bullets = [item.bullet_1, item.bullet_2, item.bullet_3, item.bullet_4, item.bullet_5]
        bullet_text = " ".join(value or "" for value in bullets)
        if len(title) < 90:
            issues.append("标题偏短，可能没有覆盖核心规格、用途和兼容词。")
        if not any(term in title.lower() for term in ["replacement", "compatible", "pack", "kit", "set"]):
            issues.append("标题缺少 replacement / compatible / pack / kit 等高意图词。")
        if sum(1 for value in bullets if value) < 4:
            issues.append("五点不完整，建议补齐规格、适配、材料、使用场景、售后说明。")
        if not any(term in bullet_text.lower() for term in ["size", "dimension", "compatible", "fit", "fits"]):
            issues.append("五点缺少尺寸/适配说明，容易影响转化和售后。")
        if not item.main_image_url:
            issues.append("缺少主图链接，建议检查图片资料是否完整。")
        audits.append({"listing": item, "issues": issues or ["基础信息完整，后续结合转化率和广告词做精细优化。"]})
    return audits


def _raw_float(row: BusinessMetric, *keys: str) -> float:
    try:
        raw = json.loads(row.raw_json or "{}")
    except json.JSONDecodeError:
        raw = {}
    for key in keys:
        value = raw.get(key)
        parsed = to_float(value)
        if parsed is not None:
            return parsed
    return 0.0


def _raw_text(row: BusinessMetric, *keys: str) -> str:
    try:
        raw = json.loads(row.raw_json or "{}")
    except json.JSONDecodeError:
        raw = {}
    for key in keys:
        value = raw.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def build_business_overview(db: Session) -> dict:
    latest_batch = (
        db.execute(
            select(ImportBatch)
            .where(ImportBatch.report_type == "business", ImportBatch.status == "success")
            .order_by(desc(ImportBatch.created_at))
            .limit(1)
        )
        .scalar_one_or_none()
    )
    if not latest_batch:
        return {"batch": None, "rows": [], "summary": {}, "insights": []}

    metrics = db.execute(select(BusinessMetric).where(BusinessMetric.batch_id == latest_batch.id)).scalars().all()
    rows = []
    for row in metrics:
        row_date = _raw_text(row, "日期", "Date")
        sku = row.sku or _raw_text(row, "SKU", "子SKU", "Child SKU", "child sku")
        asin = row.asin or _raw_text(row, "ASIN", "子ASIN", "Child ASIN", "child asin")
        title = row.title or _raw_text(row, "标题", "商品名称", "Title", "Product Name")
        sales = row.ordered_sales if row.ordered_sales is not None else _raw_float(row, "已订购商品销售额", "Ordered Product Sales")
        b2b_sales = _raw_float(row, "已订购商品销售额 - B2B", "Ordered Product Sales - B2B")
        units = row.units_ordered if row.units_ordered is not None else _raw_float(row, "已订购商品数量", "Units Ordered")
        orders = _raw_float(row, "订单商品总数", "Total Order Items")
        sessions = row.sessions if row.sessions is not None else _raw_float(row, "会话数 - 总计", "Sessions - Total", "Sessions")
        cvr = row.conversion_rate if row.conversion_rate is not None else _raw_float(row, "订单商品会话百分比", "Unit Session Percentage")
        if cvr and cvr > 1:
            cvr = cvr / 100
        asp = _raw_float(row, "平均售价", "Average Sales per Unit")
        rows.append(
            {
                "date": row_date,
                "sku": sku,
                "asin": asin,
                "title": title,
                "label": row_date or sku or asin or "",
                "dimension_type": "date" if row_date else "product",
                "sales": sales,
                "b2b_sales": b2b_sales,
                "units": units,
                "orders": orders,
                "sessions": sessions,
                "conversion_rate": cvr,
                "average_price": asp or (sales / units if units else 0),
            }
        )

    total_sales = sum(item["sales"] for item in rows)
    total_b2b = sum(item["b2b_sales"] for item in rows)
    total_units = sum(item["units"] for item in rows)
    total_orders = sum(item["orders"] for item in rows)
    total_sessions = sum(item["sessions"] for item in rows)
    record_count = len(rows)
    is_date_report = bool(rows) and sum(1 for item in rows if item["dimension_type"] == "date") >= max(1, len(rows) * 0.8)
    first_half = rows[: max(1, record_count // 2)]
    second_half = rows[max(1, record_count // 2) :]

    def avg_sales(items):
        return sum(item["sales"] for item in items) / len(items) if items else 0

    summary = {
        "day_count": record_count,
        "record_count": record_count,
        "is_date_report": is_date_report,
        "dimension_label": "日期" if is_date_report else "SKU / ASIN",
        "leaderboard_title": "天" if is_date_report else "记录",
        "detail_title": "按日明细" if is_date_report else "按 SKU/ASIN 明细",
        "average_sales_label": "日均" if is_date_report else "单记录平均",
        "total_sales": total_sales,
        "avg_daily_sales": total_sales / record_count if record_count else 0,
        "total_b2b": total_b2b,
        "b2b_ratio": total_b2b / total_sales if total_sales else 0,
        "total_units": total_units,
        "total_orders": total_orders,
        "total_sessions": total_sessions,
        "conversion_rate": total_orders / total_sessions if total_sessions else 0,
        "unit_session_rate": total_units / total_sessions if total_sessions else 0,
        "average_price": total_sales / total_units if total_units else 0,
        "average_order_value": total_sales / total_orders if total_orders else 0,
        "first_half_avg_sales": avg_sales(first_half),
        "second_half_avg_sales": avg_sales(second_half),
    }
    summary["sales_trend"] = (
        (summary["second_half_avg_sales"] - summary["first_half_avg_sales"]) / summary["first_half_avg_sales"]
        if summary["first_half_avg_sales"]
        else 0
    )
    top_days = sorted(rows, key=lambda item: item["sales"], reverse=True)[:5]
    low_days = sorted(rows, key=lambda item: item["sales"])[:5]
    insights = []
    if summary["sales_trend"] > 0.08:
        insights.append("后半段平均销售额明显高于前半段，说明近期销售动能在增强。" if is_date_report else "后半部分 SKU/ASIN 平均销售额高于前半部分，头部产品贡献更明显。")
    elif summary["sales_trend"] < -0.08:
        insights.append("后半段平均销售额低于前半段，需要检查广告、库存、价格或排名变化。" if is_date_report else "后半部分 SKU/ASIN 平均销售额低于前半部分，产品销售分层明显。")
    else:
        insights.append("整体销售较稳定，下一步重点看 SKU 级利润和广告词效率。")
    if summary["conversion_rate"] >= 0.1:
        insights.append("整体转化率超过 10%，转化基础不错，可以优先从广告放量和高利润 SKU 扩量入手。")
    else:
        insights.append("整体转化率偏低，优先排查 Listing 图片、价格、评论和流量相关性。")
    if summary["b2b_ratio"] >= 0.05:
        insights.append("B2B 销售占比不低，可以考虑数量折扣、多件装和 Business Price。")
    if summary["average_price"] >= 30:
        insights.append("平均售价较高，具备一定广告承受能力；后续要结合成本表确认真实利润。")
    return {
        "batch": latest_batch,
        "rows": rows,
        "summary": summary,
        "top_days": top_days,
        "low_days": low_days,
        "insights": insights,
    }
